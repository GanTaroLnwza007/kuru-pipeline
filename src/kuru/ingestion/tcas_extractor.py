"""TCAS structured data extractor — Gemini → TCASRecord JSON → Supabase."""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from pydantic import BaseModel, Field, field_validator
from tenacity import retry, retry_if_exception, stop_after_attempt, wait_exponential

from kuru.db import supabase_client as db
from kuru.ingestion.text_extractor import extract_text, full_text
from kuru.ingestion.utils import is_transient_error, safe_print
from kuru.llm import LLM_MODEL, get_client

# Maximum sheets processed per xlsx workbook — guard against runaway API costs.
MAX_XLSX_SHEETS = 20


# ─────────────────────────────────────────
# Schema
# ─────────────────────────────────────────

class TCASRecord(BaseModel):
    program_name_raw: str = Field(description="Program name as extracted from the PDF")
    faculty: str = Field(default="")
    round: str = Field(description="TCAS round: 'round1' | 'round2' | 'round3' | 'round4'")
    quota: int | None = Field(default=None)
    gpax_min: float | None = Field(default=None)
    exam_criteria: dict[str, Any] | None = Field(default=None)
    portfolio_requirements: dict[str, Any] | None = Field(default=None)
    deadlines: dict[str, Any] | None = Field(default=None)

    @field_validator("gpax_min")
    @classmethod
    def _gpax_must_be_on_4_scale(cls, v: float | None) -> float | None:
        # GPAX is 0.0–4.0. Values > 4.0 mean the extractor confused a score weight
        # percentage (e.g. "GPAX 20%") with the minimum GPA — discard them.
        if v is not None and v > 4.0:
            return None
        return v


EXTRACTION_PROMPT = """You are a structured data extractor for Thai university admission documents.

Extract ALL program admission records from the following TCAS document text.
Return a JSON array where each element is an admission record with these fields:
- program_name_raw: exact program name in Thai as written in the document
- faculty: faculty/department name
- round: "round1", "round2", "round3", or "round4"
- quota: number of available seats (integer, null if not found)
- gpax_min: minimum cumulative GPA (GPAX) on a 4.0 scale (e.g. 2.75, 3.00, 3.25). Must be between 0.0 and 4.0. If GPAX appears as a score weight percentage (e.g. "GPAX 20%"), that is a weight — NOT a minimum GPA, leave null.
- exam_criteria: object with exam names as keys and percentage weights (0–100), e.g. {"TGAT": {"weight": 30}, "TPAT3": {"weight": 70}}
- portfolio_requirements: object describing portfolio items required
- deadlines: object with date keys, e.g. {"apply_start": "2025-10-01", "apply_end": "2025-10-10"}

Output ONLY valid JSON — no markdown fences, no commentary.

Document text:
{text}
"""


# ─────────────────────────────────────────
# Extraction
# ─────────────────────────────────────────

@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=2, min=15, max=120), retry=retry_if_exception(is_transient_error), reraise=True)
def _call_llm(text: str) -> str:
    response = get_client().chat.completions.create(
        model=LLM_MODEL,
        messages=[{"role": "user", "content": EXTRACTION_PROMPT.replace("{text}", text[:200000])}],
        temperature=0.0,
    )
    return response.choices[0].message.content or "[]"


def _parse_records(raw_json: str) -> list[dict[str, Any]]:
    # Strip markdown fences if Gemini ignores the instruction
    cleaned = re.sub(r"```(?:json)?|```", "", raw_json).strip()
    try:
        data = json.loads(cleaned)
    except json.JSONDecodeError as exc:
        safe_print(f"  JSON parse error: {exc}\n  Raw response (first 200 chars): {cleaned[:200]}")
        return []
    if isinstance(data, dict):
        # Sometimes Gemini wraps in {"records": [...]}
        data = data.get("records", data.get("data", [data]))
    if not isinstance(data, list):
        safe_print(f"  Unexpected Gemini response type ({type(data).__name__}), value: {repr(cleaned[:200])}")
        return []
    return data


def _build_records(raw_records: list[dict[str, Any]]) -> list[TCASRecord]:
    records: list[TCASRecord] = []
    for r in raw_records:
        try:
            records.append(TCASRecord(**r))
        except Exception as exc:
            safe_print(f"  Warning: skipping malformed record — {exc}: {r}")
    return records


def extract_tcas_from_pdf(
    pdf_path: str | Path,
    verbose: bool = False,
) -> list[TCASRecord]:
    """Extract structured TCAS records from a PDF file."""
    pdf_path = Path(pdf_path)
    if verbose:
        safe_print(f"Extracting text from {pdf_path.name} …")

    pages = extract_text(pdf_path, use_vision_fallback=True, verbose=verbose)
    doc_text = full_text(pages)

    if verbose:
        safe_print(f"  Extracted {len(doc_text)} chars. Calling Gemini for structured extraction …")

    return _build_records(_parse_records(_call_llm(doc_text)))


def _sheet_to_text(ws: Any) -> str:
    rows = []
    for row in ws.iter_rows(values_only=True):
        row_text = "\t".join("" if v is None else str(v) for v in row)
        if row_text.strip():
            rows.append(row_text)
    return "\n".join(rows)


def extract_tcas_from_xlsx(
    xlsx_path: str | Path,
    verbose: bool = False,
) -> list[TCASRecord]:
    """Extract structured TCAS records from an xlsx spreadsheet.

    Each sheet is processed as a separate Gemini call. Duplicate
    (program_name_raw, round) pairs across sheets are deduplicated (first wins).
    """
    import openpyxl

    xlsx_path = Path(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    sheets = wb.sheetnames[:MAX_XLSX_SHEETS]
    if len(wb.sheetnames) > MAX_XLSX_SHEETS:
        safe_print(f"  Warning: workbook has {len(wb.sheetnames)} sheets; processing first {MAX_XLSX_SHEETS} only.")

    if verbose:
        safe_print(f"  xlsx: {len(sheets)} sheet(s) — processing each separately …")

    all_records: list[TCASRecord] = []
    seen: set[str] = set()

    for sheet_name in sheets:
        ws = wb[sheet_name]
        text = f"=== Sheet: {sheet_name} ===\n{_sheet_to_text(ws)}"

        if verbose:
            safe_print(f"    Sheet '{sheet_name}': {len(text)} chars -> Gemini ...")

        sheet_new = 0
        for rec in _build_records(_parse_records(_call_llm(text))):
            key = f"{rec.program_name_raw}|{rec.round}"
            if key not in seen:
                seen.add(key)
                all_records.append(rec)
                sheet_new += 1

        if verbose:
            safe_print(f"    -> {sheet_new} new record(s) (total so far: {len(all_records)})")

    if verbose:
        safe_print(f"  xlsx done: {len(all_records)} unique TCAS records extracted.")

    return all_records


def store_tcas_records(
    records: list[TCASRecord],
    source_file: str,
) -> None:
    """Upsert TCASRecord objects to Supabase."""
    client = db.get_client()
    rows = [
        {**r.model_dump(), "source_file": source_file}
        for r in records
    ]
    db.upsert_tcas_records(client, rows)
